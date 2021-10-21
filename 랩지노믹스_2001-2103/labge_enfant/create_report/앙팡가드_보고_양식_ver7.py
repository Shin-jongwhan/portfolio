# 보고 양식(내부) for windows OS
################# 설명 ##################
# 1. sLoad_workbook_dir 에 "C:\\Users\\user\\Desktop\\신종환_앙팡가드\\신종환_200106~\\2020_앙팡가드_결과파일_1월_ori.xlsx" 와 같이 본인의 엑셀 경로 넣으시면 됩니다.
# 2. fWrite.write("신종환 올림.") >> 본인 이름으로 고쳐쓰기
# 3. 터미널 키고
# > python [앙팡가드_내부보고_약식.py] -s [sheet 이름] 적으시면 엑셀 경로에 텍스트 파일 생성됩니다.
# 양성 - 양성 의심 - LR - SR - 성별 확인 - 음성(benign) - 음성 순
################# 설명 ##################

import sys, getopt
from openpyxl import load_workbook

def option() :
    if __name__=='__main__':
        print(sys.argv)
        optlist, args = getopt.getopt(sys.argv[1:], 's:')
        print("Description : Option (-s file)")
        print()
        print(optlist)
        print()
    
    return optlist


def option2() :
    if __name__=='__main__':
        optlist, args = getopt.getopt(sys.argv[1:], 's:')
    
    return optlist


def report_parents(intReportNum, lsResult, Load_ws, fWrite) :
    nParent_test_num = 0
    boolParent_test = False
    for index in range(0, len(lsResult)) :
        for i in range(1, len(lsResult[index])) :
            if "Father" in Load_ws.cell(lsResult[index][i], 8).value or "Mother" in Load_ws.cell(lsResult[index][i], 8).value :
                nParent_test_num += 1
                boolParent_test = True
    if nParent_test_num != 0 :
        intReportNum += 1
        fWrite.write("{0}. 부모 검사 {1}건(같은 아기의 부모검사가 2개 인 경우 1건으로 고치기)\n".format( intReportNum, nParent_test_num ))
    for index in range(0, len(lsResult)) :
        for i in range(1, len(lsResult[index])) :
            if "Father" in Load_ws.cell(lsResult[index][i], 8).value or "Mother" in Load_ws.cell(lsResult[index][i], 8).value :
                #print(Load_ws.cell(lsResult[index][i], 8).value)
                if index == 0 :
                    fWrite.write("  - {0}\t{1}_{2}\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value, Load_ws.cell(lsResult[index][i], 8).value ))
                elif index == 1 :
                    fWrite.write("  - {0}\t{1}_{2}\t{3}\t{4}\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value, Load_ws.cell(lsResult[index][i], 8).value, Load_ws.cell(lsResult[index][i], 16).value, Load_ws.cell(lsResult[index][i], 17).value ))
                elif index == 2 :
                    fWrite.write("  - {0}\t{1}_{2}\t{3}\t{4}\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value, Load_ws.cell(lsResult[index][i], 8).value, Load_ws.cell(lsResult[index][i], 16).value, Load_ws.cell(lsResult[index][i], 17).value ))
                elif index == 3 :
                    fWrite.write("  - {0}\t{1}_{2}\t{3}\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value, Load_ws.cell(lsResult[index][i], 8).value, Load_ws.cell(lsResult[index][i], 16).value ))
                elif index == 4 :
                    fWrite.write("  - {0}\t{1}_{2}\t{3}\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value, Load_ws.cell(lsResult[index][i], 8).value, Load_ws.cell(lsResult[index][i], 16).value ))
                elif index == 5 :
                    fWrite.write("  - {0}\t{1}_{2}\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value, Load_ws.cell(lsResult[index][i], 8).value ))
                elif index == 6 :
                    fWrite.write("  - {0}\t{1}_{2}\t{3}\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value, Load_ws.cell(lsResult[index][i], 8).value, Load_ws.cell(lsResult[index][i], 20).value ))
                elif index == 7 :
                    fWrite.write("  - {0}\t{1}_{2}\t음성\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value, Load_ws.cell(lsResult[index][i], 8).value))
                elif index == 8 :
                    fWrite.write("  - {0}\t{1}_{2}\t{3}\t{4}\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value, Load_ws.cell(lsResult[index][i], 8).value, Load_ws.cell(lsResult[index][i], 16).value, Load_ws.cell(lsResult[index][i], 17).value ))
    if boolParent_test == True :
        fWrite.write("\n")
    for index in range(0, len(lsResult)) :
        for i in range(len(lsResult[index]) - 1, 0, -1) :
            if "Father" in Load_ws.cell(lsResult[index][i], 8).value or "Mother" in Load_ws.cell(lsResult[index][i], 8).value :
                del lsResult[index][i]
                lsResult[index][0] -= 1

    return nParent_test_num, intReportNum, boolParent_test


#def request_review(intReportNum, lsResult, Load_ws, fWrite)


def report(intReportNum, index, lsResult, Load_ws, fWrite) :
    if index == 0 :
        fWrite.write("{0}. 리뷰 진행중 {1}건\n".format( intReportNum, lsResult[index][0] ))
        for i in range(1, len(lsResult[index])) :
            if i != len(lsResult[index]) - 1 : 
                fWrite.write("  - {0}\t{1}\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value ))
            else :
                fWrite.write("  - {0}\t{1}\n\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value ))
    if index == 10 :
        fWrite.write("{0}. 의견 요청 {1}건\n".format( intReportNum, lsResult[index][0] ))
        for i in range(1, len(lsResult[index])) :
            if i != len(lsResult[index]) - 1 : 
                fWrite.write("  - {0}\t{1}\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value ))
            else :
                fWrite.write("  - {0}\t{1}\n\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value ))
    elif index == 1 :
        fWrite.write("{0}. 양성 {1}건\n".format( intReportNum, lsResult[index][0] ))
        for i in range(1, len(lsResult[index])) :
            if i != len(lsResult[index]) - 1 : 
                fWrite.write("  - {0}\t{1}\t{2}\t{3}\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value, Load_ws.cell(lsResult[index][i], 16).value, Load_ws.cell(lsResult[index][i], 17).value ))
            else :
                fWrite.write("  - {0}\t{1}\t{2}\t{3}\n\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value, Load_ws.cell(lsResult[index][i], 16).value, Load_ws.cell(lsResult[index][i], 17).value ))
    elif index == 2 :
        fWrite.write("{0}. 양성 의심(재검) {1}건\n".format( intReportNum, lsResult[index][0] ))
        for i in range(1, len(lsResult[index])) :
            if i != len(lsResult[index]) - 1 : 
                fWrite.write("  - {0}\t{1}\t{2}\t{3}\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value, Load_ws.cell(lsResult[index][i], 16).value, Load_ws.cell(lsResult[index][i], 17).value ))
            else :
                fWrite.write("  - {0}\t{1}\t{2}\t{3}\n\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value, Load_ws.cell(lsResult[index][i], 16).value, Load_ws.cell(lsResult[index][i], 17).value ))
    elif index == 3 :
        fWrite.write("{0}. LR 재검 {1}건\n".format( intReportNum, lsResult[index][0] ))
        for i in range(1, len(lsResult[index])) :
            if i != len(lsResult[index]) - 1 : 
                fWrite.write("  - {0}\t{1}\t{2}\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value, Load_ws.cell(lsResult[index][i], 16).value ))
            else :
                fWrite.write("  - {0}\t{1}\t{2}\n\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value, Load_ws.cell(lsResult[index][i], 16).value ))
    elif index == 4 :
        fWrite.write("{0}. SR 재검 {1}건\n".format( intReportNum, lsResult[index][0] ))
        for i in range(1, len(lsResult[index])) :
            if i != len(lsResult[index]) - 1 : 
                fWrite.write("  - {0}\t{1}\t{2}\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value, Load_ws.cell(lsResult[index][i], 16).value ))
            else :
                fWrite.write("  - {0}\t{1}\t{2}\n\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value, Load_ws.cell(lsResult[index][i], 16).value ))
    elif index == 5 :
        fWrite.write("{0}. 성별 확인중 {1}건\n".format( intReportNum, lsResult[index][0] ))
        for i in range(1, len(lsResult[index])) :
            if i != len(lsResult[index]) - 1 : 
                fWrite.write("  - {0}\t{1}\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value ))
            else :
                fWrite.write("  - {0}\t{1}\n\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value ))
    elif index == 6 :
        fWrite.write("{0}. 재채혈 요청 및 재검 {1}건\n".format( intReportNum, lsResult[index][0] ))
        for i in range(1, len(lsResult[index])) :
            if i != len(lsResult[index]) - 1 : 
                fWrite.write("  - {0}\t{1}\t{2}\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value, Load_ws.cell(lsResult[index][i], 20).value ))
            else :
                fWrite.write("  - {0}\t{1}\t{2}\n\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value, Load_ws.cell(lsResult[index][i], 20).value ))
    elif index == 7 :
        fWrite.write("{0}. 성별 확인 재검 {1}건\n".format( intReportNum, lsResult[index][0] ))
        for i in range(1, len(lsResult[index])) :
            if i != len(lsResult[index]) - 1 : 
                fWrite.write("  - {0}\t{1}\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value ))
            else :
                fWrite.write("  - {0}\t{1}\n\n".format( Load_ws.cell(lsResult[index][i], 2).value.split("-")[0], Load_ws.cell(lsResult[index][i], 10).value ))
    

def write_clinvar(lsResult, Load_ws, fWrite) :
    fWrite.write("ClinVar\n")
    



def inner_part_report() :
    # 값이 없으면 None 출력한다.
    # load_ws.cell(column, row)
    # row (i, 1) RunID, (i, 2) Sample ID, (i, 10) Initial, (i, 16) 음성/양성, (i, 17) EnfantGuard Disease, (i, 20) issue
    optlist = option()

    # "C:\\Users\\user\\Desktop\\신종환_앙팡가드\\신종환_200106~\\2020_앙팡가드_결과파일_1월_ori.xlsx"
    # sLoad_workbook_dir = optlist[0][1]
    sLoad_workbook_dir = "C:\\Users\\user\\Dropbox\\백업\앙팡_백업\\앙팡가드_결과파일\\2020_앙팡가드_결과파일_201019.xlsx"
    
    load_wb = load_workbook(sLoad_workbook_dir, data_only = True)

    load_ws = load_wb[ optlist[0][1] ]
    sSheet = optlist[0][1]
    # sSheet = "20200110"
    # load_ws = load_wb[sSheet]

    # 셀 주소로 값 출력
    print(load_ws['Q2'].value)

    # 참고
    #if load_ws['Q2'].value == None :
    #    print("True")

    # 셀 좌표로 값 출력
    print(load_ws.cell(1,1).value)

    intRowEnd = 0
    for i in range(1, 10000) :      # find column end
        if load_ws.cell(i, 1).value == None :
            intRowEnd = i - 1
            print(intRowEnd)
            break

    # [0] 리뷰 진행 중, [1] 양성 [2] 양성 의심, [3] LR, [4] SR [5] 성별 확인중 [6] 재채혈 [7] 성별 확인 재검 [8] 음성(benign) [9] 음성
    print("[0] 리뷰 진행 중, [1] 양성 [2] 양성 의심, [3] LR, [4] SR [5] 성별 확인중")
    print("[6] 재채혈 [7] 성별 확인 재검 [8] 음성(benign) [9] 음성 [10] 의견 요청")
    lsResult = [ [0], [0], [0], [0], [0], [0], [0], [0], [0], [0], [0] ]
    intReportNum = 0
    for i in range(2, intRowEnd + 1) :
        if load_ws.cell(i, 16).value != None :
            if "리뷰" in load_ws.cell(i, 16).value :
                lsResult[0][0] += 1
                lsResult[0].append(i)
            elif "양성" in load_ws.cell(i, 16).value and "의심" not in load_ws.cell(i, 16).value and "성별" not in load_ws.cell(i, 16).value and "리뷰" not in load_ws.cell(i, 16).value :
                lsResult[1][0] += 1
                lsResult[1].append(i)
            elif "양성" in load_ws.cell(i, 16).value and "의심" in load_ws.cell(i, 16).value and "성별" not in load_ws.cell(i, 16).value and "리뷰" not in load_ws.cell(i, 16).value :
                lsResult[2][0] += 1
                lsResult[2].append(i)
            elif "LR" in load_ws.cell(i, 16).value and "성별" not in load_ws.cell(i, 16).value and "리뷰" not in load_ws.cell(i, 16).value :
                lsResult[3][0] += 1
                lsResult[3].append(i)
            elif "SR" in load_ws.cell(i, 16).value and "성별" not in load_ws.cell(i, 16).value and "리뷰" not in load_ws.cell(i, 16).value :
                lsResult[4][0] += 1
                lsResult[4].append(i)
            elif "성별확인" in load_ws.cell(i, 16).value or "성별 확인" in load_ws.cell(i, 16).value and "재검" not in load_ws.cell(i, 16).value and "리뷰" not in load_ws.cell(i, 16).value :
                lsResult[5][0] += 1
                lsResult[5].append(i)
            elif "재채혈" in load_ws.cell(i, 16).value :
                lsResult[6][0] += 1
                lsResult[6].append(i)
            elif "성별확인" in load_ws.cell(i, 16).value or "성별 확인" in load_ws.cell(i, 16).value and "재검" in load_ws.cell(i, 16).value and "리뷰" not in load_ws.cell(i, 16).value :
                lsResult[7][0] += 1
                lsResult[7].append(i)
            elif "음성" in load_ws.cell(i, 16).value and "benign" in load_ws.cell(i, 16).value and "성별" not in load_ws.cell(i, 16).value and "리뷰" not in load_ws.cell(i, 16).value:
                lsResult[8][0] += 1
                lsResult[8].append(i)
            elif "음성" in load_ws.cell(i, 16).value and "benign" not in load_ws.cell(i, 16).value and "성별" not in load_ws.cell(i, 16).value and "리뷰" not in load_ws.cell(i, 16).value :
                lsResult[9][0] += 1
                lsResult[9].append(i)
            elif "의견" in load_ws.cell(i, 16).value :
                lsResult[10][0] += 1
                lsResult[10].append(i)
            
    intNumofSampleEntry = 0
    for i in range(0, len(lsResult)) :
        intNumofSampleEntry += lsResult[i][0]

    print("lsResult : ", lsResult)
    print("intNumofSampleEntry : ",intNumofSampleEntry)
    
    sDate = load_ws.cell(2,1).value.split("_")[0][2:4] + "/" + load_ws.cell(2,1).value.split("_")[0][4:6]
    print(sDate)

    fWrite = open("\\".join( sLoad_workbook_dir.split("\\")[:-1] ) + "\\" + sSheet + "_보고양식.txt", 'w', -1, "utf-8")
    fWrite.write("안녕하세요.\n\n")
    fWrite.write("앙팡가드 {0}일자 {1}건 보고드립니다.\n\n".format(sDate, str(intNumofSampleEntry)))

    ############################################################
    # report
    intReportNum = 0
    boolParent_test = False
    nParent_test_num = 0
    nParent_test_num, intReportNum, boolParent_test = report_parents(intReportNum, lsResult, load_ws, fWrite)
    for i in range(0, len(lsResult)) :
        if i != 8 or i != 9 : 
            if lsResult[i][0] != 0 :
                print(i, lsResult[i])
                intReportNum += 1
                report(intReportNum, i, lsResult, load_ws, fWrite)
        if i == len(lsResult) - 1 : 
            if intReportNum == 0 :
                fWrite.write("모두 음성 판정하였습니다.\n\n")
                break
            else :
                if lsResult[8][0] + lsResult[9][0] != 0 : 
                    fWrite.write("외 {0}건 음성 판정하였습니다.\n\n".format(str(lsResult[8][0] + lsResult[9][0])))
                break

    ############################################################
    # 장애유 판정(음성(benign) 이거나 음성일 경우)
    nDisorder_line = 0
    nDisorder_entry = 0
    boolDisorder_header_line = True
    for i in range(0, len(lsResult)) :
        if i < 7 : 
            continue
        if len(lsResult[i]) == 1 :
            continue
        for j in range(1, len(lsResult[i])) :
            if load_ws.cell(lsResult[i][j], 20).value != None :
                if "장애" in load_ws.cell(lsResult[i][j], 20).value :
                    nDisorder_entry += 1
    for i in range(0, len(lsResult)) :
        if i >= 7 :
            if boolDisorder_header_line == True and nDisorder_entry != 0 :
                fWrite.write("* 장애유 {0}건 -> 음성\n".format(nDisorder_entry))
                boolDisorder_header_line = False
            for j in range(1, len(lsResult[i])) :
                if load_ws.cell(lsResult[i][j], 20).value != None and "장애" in load_ws.cell(lsResult[i][j], 20).value :
                    fWrite.write("  - " + load_ws.cell(lsResult[i][j], 2).value.split("-")[0] + "\t")
                    fWrite.write(load_ws.cell(lsResult[i][j], 10).value + "\n")
                    #fWrite.write(load_ws.cell(lsResult[i][j], 10).value + "\t")
                    #fWrite.write("장애유 -> 음성 판정\n")
                    nDisorder_line += 1
                if i == len(lsResult) - 1 and j == len(lsResult[i]) - 1 and nDisorder_line != 0 :
                    fWrite.write("\n")
    ############################################################
    # ClinVar
    # load_ws.cell(lsResult[i][j], 19).value = ClinVar
    #fWrite.write("ClinVar\n")
    #for i in range(0, len(lsResult)) :
    #    for j in range(1, len(lsResult[i])) :
    #        if load_ws.cell(lsResult[i][j], 19).value != None :
    #            fWrite.write("  - " + load_ws.cell(lsResult[i][j], 2).value.split("-")[0] + "\t")
    #            fWrite.write(load_ws.cell(lsResult[i][j], 10).value + "\t")
    #            fWrite.write(load_ws.cell(lsResult[i][j], 19).value + "\n")
    #        if i == len(lsResult) - 1 and j == len(lsResult[i]) - 1 :
    #            fWrite.write("\n")
        
        

    fWrite.write("검토 부탁 드립니다. \n\n")
    fWrite.write("감사합니다.\n\n")
    fWrite.write("신종환 올림.\n")
    fWrite.write("===============================================\n")
    
    print("len(lsReslt)", len(lsResult))        # for check
    for i in range(0, len(lsResult)) :
        print("len(lsResult[i]) ", len(lsResult[i]))        # for check
        if len(lsResult[i]) > 1 :
            for j in range(1, len(lsResult[i])) :
                #print(i, j)
                if 0 <= i and i <= 7 :
                    if load_ws.cell(lsResult[i][j], 20).value == None :
                        fWrite.write(load_ws.cell(lsResult[i][j], 2).value.split("-")[0] + "\t")
                        fWrite.write(load_ws.cell(lsResult[i][j], 10).value + "\t")
                        fWrite.write(load_ws.cell(lsResult[i][j], 16).value + "\n")
                    else :
                        fWrite.write(load_ws.cell(lsResult[i][j], 2).value.split("-")[0] + "\t")
                        fWrite.write(load_ws.cell(lsResult[i][j], 10).value + "\t")
                        fWrite.write(load_ws.cell(lsResult[i][j], 16).value + "\t")
                        fWrite.write(load_ws.cell(lsResult[i][j], 20).value + "\n")
                #elif i == 6 :
                    #if load_ws.cell(lsResult[i][j], 20).value == None : 
                    #    fWrite.write(load_ws.cell(lsResult[i][j], 2).value.split("-")[0] + "\t")
                    #    fWrite.write(load_ws.cell(lsResult[i][j], 10).value + "\t")
                    #    fWrite.write(load_ws.cell(lsResult[i][j], 16).value + "\n")
                    #elif load_ws.cell(lsResult[i][j], 20).value != None and "장애" not in load_ws.cell(lsResult[i][j], 20).value : 
                    #    fWrite.write(load_ws.cell(lsResult[i][j], 2).value.split("-")[0] + "\t")
                    #    fWrite.write(load_ws.cell(lsResult[i][j], 10).value + "\t")
                    #    fWrite.write(load_ws.cell(lsResult[i][j], 16).value + "\t")
                    #    fWrite.write(load_ws.cell(lsResult[i][j], 20).value + "\n")
                    #elif load_ws.cell(lsResult[i][j], 20).value != None and "장애" in load_ws.cell(lsResult[i][j], 20).value :
                    #    fWrite.write(load_ws.cell(lsResult[i][j], 2).value.split("-")[0] + "\t")
                    #    fWrite.write(load_ws.cell(lsResult[i][j], 10).value + "\t")
                    #    fWrite.write("장애유 -> 음성(benign) 판정\t")
                    #    fWrite.write(load_ws.cell(lsResult[i][j], 20).value + "\n")
                #elif i == 7 :       # write first if issue exist
                    #if j != len(lsResult[i]) - 1 and load_ws.cell(lsResult[i][j], 20).value != None and "장애" not in load_ws.cell(lsResult[i][j], 20).value :
                    #    fWrite.write(load_ws.cell(lsResult[i][j], 2).value.split("-")[0] + "\t")
                    #    fWrite.write(load_ws.cell(lsResult[i][j], 10).value + "\t")
                    #    fWrite.write(load_ws.cell(lsResult[i][j], 20).value + "\n")
                    #elif j != len(lsResult[i]) - 1 and load_ws.cell(lsResult[i][j], 20).value != None and "장애" in load_ws.cell(lsResult[i][j], 20).value :
                    #    fWrite.write(load_ws.cell(lsResult[i][j], 2).value.split("-")[0] + "\t")
                    #    fWrite.write(load_ws.cell(lsResult[i][j], 10).value + "\t")
                    #    fWrite.write("장애유 -> 음성 판정\n")
                    #elif j == len(lsResult[i]) - 1 and load_ws.cell(lsResult[i][j], 20).value != None and "장애" not in load_ws.cell(lsResult[i][j], 20).value :
                    #    fWrite.write(load_ws.cell(lsResult[i][j], 2).value.split("-")[0] + "\t")
                    #    fWrite.write(load_ws.cell(lsResult[i][j], 10).value + "\t")
                    #    fWrite.write(load_ws.cell(lsResult[i][j], 20).value)
                    #elif j == len(lsResult[i]) - 1 and load_ws.cell(lsResult[i][j], 20).value != None and "장애" in load_ws.cell(lsResult[i][j], 20).value :
                    #    fWrite.write(load_ws.cell(lsResult[i][j], 2).value.split("-")[0] + "\t")
                    #    fWrite.write(load_ws.cell(lsResult[i][j], 10).value + "\t")
                    #    fWrite.write("장애유 -> 음성 판정")
    # for loop end
                
    fWrite.close()
    load_wb.close()


def final_report() :
    # 값이 없으면 None 출력한다.
    # load_ws.cell(column, row)
    # row (i, 1) RunID, (i, 2) Sample ID, (i, 10) Initial, (i, 16) 음성/양성, (i, 17) EnfantGuard Disease, (i, 20) issue
    optlist = option2()

    # "C:\\Users\\user\\Desktop\\신종환_앙팡가드\\신종환_200106~\\2020_앙팡가드_결과파일_1월_ori.xlsx"
    # sLoad_workbook_dir = optlist[0][1]
    sLoad_workbook_dir = "C:\\Users\\user\\Dropbox\\백업\앙팡_백업\\앙팡가드_결과파일\\2020_앙팡가드_결과파일_201019.xlsx"
    
    load_wb = load_workbook(sLoad_workbook_dir, data_only = True)

    load_ws = load_wb[ optlist[0][1] ]
    sSheet = optlist[0][1]
    # sSheet = "20200110"
    # load_ws = load_wb[sSheet]

    # 셀 주소로 값 출력
    #print(load_ws['Q2'].value)

    # 참고
    #if load_ws['Q2'].value == None :
    #    print("True")

    # 셀 좌표로 값 출력
    #print(load_ws.cell(1,1).value)

    intRowEnd = 0
    for i in range(1, 10000) :      # find column end
        if load_ws.cell(i, 1).value == None :
            intRowEnd = i - 1
            #print(intRowEnd)
            break

    # [0] 리뷰 진행 중, [1] 양성 [2] 양성 의심, [3] LR, [4] SR [5] 성별 확인 [6] 재채혈 [7] 성별 확인 재검 [8] 음성(benign) [9] 음성
    lsResult = [ [0], [0], [0], [0], [0], [0], [0], [0], [0], [0] ]
    intReportNum = 0
    for i in range(2, intRowEnd + 1) :
        if load_ws.cell(i, 16).value != None :
            if "리뷰" in load_ws.cell(i, 16).value :
                lsResult[0][0] += 1
                lsResult[0].append(i)
            elif "양성" in load_ws.cell(i, 16).value and "의심" not in load_ws.cell(i, 16).value and "성별" not in load_ws.cell(i, 16).value and "리뷰" not in load_ws.cell(i, 16).value :
                lsResult[1][0] += 1
                lsResult[1].append(i)
            elif "양성" in load_ws.cell(i, 16).value and "의심" in load_ws.cell(i, 16).value and "성별" not in load_ws.cell(i, 16).value and "리뷰" not in load_ws.cell(i, 16).value :
                lsResult[2][0] += 1
                lsResult[2].append(i)
            elif "LR" in load_ws.cell(i, 16).value and "성별" not in load_ws.cell(i, 16).value and "리뷰" not in load_ws.cell(i, 16).value :
                lsResult[3][0] += 1
                lsResult[3].append(i)
            elif "SR" in load_ws.cell(i, 16).value and "성별" not in load_ws.cell(i, 16).value and "리뷰" not in load_ws.cell(i, 16).value :
                lsResult[4][0] += 1
                lsResult[4].append(i)
            elif "성별확인" in load_ws.cell(i, 16).value or "성별 확인" in load_ws.cell(i, 16).value and "재검" not in load_ws.cell(i, 16).value and "리뷰" not in load_ws.cell(i, 16).value :
                lsResult[5][0] += 1
                lsResult[5].append(i)
            elif "재채혈" in load_ws.cell(i, 16).value :
                lsResult[6][0] += 1
                lsResult[6].append(i)
            elif "성별확인" in load_ws.cell(i, 16).value or "성별 확인" in load_ws.cell(i, 16).value and "재검" in load_ws.cell(i, 16).value and "리뷰" not in load_ws.cell(i, 16).value :
                lsResult[7][0] += 1
                lsResult[7].append(i)
            elif "음성" in load_ws.cell(i, 16).value and "benign" in load_ws.cell(i, 16).value and "성별" not in load_ws.cell(i, 16).value and "리뷰" not in load_ws.cell(i, 16).value:
                lsResult[8][0] += 1
                lsResult[8].append(i)
            elif "음성" in load_ws.cell(i, 16).value and "benign" not in load_ws.cell(i, 16).value and "성별" not in load_ws.cell(i, 16).value and "리뷰" not in load_ws.cell(i, 16).value :
                lsResult[9][0] += 1
                lsResult[9].append(i)
            
    intNumofSampleEntry = 0
    for i in range(0, len(lsResult)) :
        intNumofSampleEntry += lsResult[i][0]

    #print("lsResult : ", lsResult)
    #print("intNumofSampleEntry : ",intNumofSampleEntry)
    
    sDate = load_ws.cell(2,1).value.split("_")[0][2:4] + "/" + load_ws.cell(2,1).value.split("_")[0][4:6]
    #print(sDate)

    fWrite = open("\\".join( sLoad_workbook_dir.split("\\")[:-1] ) + "\\" + sSheet + "_최종보고양식.txt", 'w', -1, "utf-8")
    fWrite.write("안녕하세요.\n\n")
    fWrite.write("앙팡가드 {0}일자 {1}건 보고드립니다.\n\n".format(sDate, str(intNumofSampleEntry)))

    intReportNum = 0
    boolParent_test = False
    nParent_test_num = 0
    nParent_test_num, intReportNum, boolParent_test = report_parents(intReportNum, lsResult, load_ws, fWrite)
    for i in range(0, len(lsResult)) :
        if i < len(lsResult) - 2 :
            if lsResult[i][0] != 0 :
                intReportNum += 1
                report(intReportNum, i, lsResult, load_ws, fWrite)
        else :
            if intReportNum == 0 :
                fWrite.write("모두 음성 판정하였습니다.\n\n")
                break
            else :
                if lsResult[8][0] + lsResult[9][0] != 0 : 
                    fWrite.write("외 {0}건 음성 판정하였습니다.\n\n".format(str(lsResult[8][0] + lsResult[9][0])))
                break

    #fWrite.write("검토 부탁 드립니다. \n\n")
    fWrite.write("감사합니다.\n\n")
    #fWrite.write("신종환 드림.\n")
    fWrite.write("===============================================\n")
    
    #print("len(lsReslt)", len(lsResult))        # for check
    for i in range(0, len(lsResult)) :
        #print("len(lsResult[i]) ", len(lsResult[i]))        # for check
        if len(lsResult[i]) > 1 :
            for j in range(1, len(lsResult[i])) :
                #print(i, j)
                if 0 <= i and i <= 6 :
                    if i == 3 or i == 4:
                        fWrite.write(load_ws.cell(lsResult[i][j], 2).value.split("-")[0] + "\t")
                        fWrite.write(load_ws.cell(lsResult[i][j], 10).value + "\t")
                        fWrite.write(load_ws.cell(lsResult[i][j], 16).value.split("(")[0] + "\n")
                    else : 
                        fWrite.write(load_ws.cell(lsResult[i][j], 2).value.split("-")[0] + "\t")
                        fWrite.write(load_ws.cell(lsResult[i][j], 10).value + "\t")
                        fWrite.write(load_ws.cell(lsResult[i][j], 16).value + "\n")
                elif i == 8 :
                    fWrite.write(load_ws.cell(lsResult[i][j], 2).value.split("-")[0] + "\t")
                    fWrite.write(load_ws.cell(lsResult[i][j], 10).value + "\n")
                elif i == 9 :       # write first if issue exist
                    if j != len(lsResult[i]) - 1 :
                        fWrite.write(load_ws.cell(lsResult[i][j], 2).value.split("-")[0] + "\t")
                        fWrite.write(load_ws.cell(lsResult[i][j], 10).value + "\n")
                    else :
                        fWrite.write(load_ws.cell(lsResult[i][j], 2).value.split("-")[0] + "\t")
                        fWrite.write(load_ws.cell(lsResult[i][j], 10).value)
    # for loop end
                
    fWrite.close()
    load_wb.close()
    
          
def main() :
    inner_part_report()
    final_report()


main()
